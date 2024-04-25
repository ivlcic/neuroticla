import os
import os
import sys
from datetime import datetime
from typing import List, Dict, Optional

import networkx as nx
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell, WriteOnlyCell
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill
from sklearn.metrics.pairwise import cosine_similarity

from ..esdl.article import Article


def cluster_louvain(articles: List[Article], embed_field_name: str, similarity_threshold: float = 0.84):
    embeddings = []
    [embeddings.append(x.data[embed_field_name]) for x in articles if embed_field_name in x.data]
    embeddings = np.array(embeddings)
    labels = [0] * len(embeddings)
    x = cosine_similarity(embeddings, embeddings)
    similarity_matrix = x > similarity_threshold
    G = nx.from_numpy_array(similarity_matrix)
    communities = nx.algorithms.community.louvain_communities(G, resolution=0.1)
    for community in communities:
        initial_member = min(community)
        for member in community:
            labels[member] = initial_member

    clusters = {}
    for a, lbl in zip(articles, labels):
        if lbl not in clusters:
            clusters[lbl] = [a]
        else:
            clusters[lbl].append(a)
    clusters = dict(sorted(clusters.items(), key=lambda x: -len(x[1])))
    consistent = {}
    for k in clusters.keys():
        c_articles: List[Article] = clusters[k]
        c_articles.sort(key=lambda article: (article.mediaReach, article.created), reverse=True)
        consistent[c_articles[0].uuid] = c_articles
    return consistent


def cluster_print(clusters: Dict[int, List[Article]], file_name: Optional[str] = None):
    with open(file_name, 'w', encoding='utf-8') if file_name else sys.stdout as output:
        for k in clusters.keys():
            articles: List[Article] = clusters[k]
            print(f'Cluster [{articles[0].title}]', file=output)
            for x, a in enumerate(articles):
                if x == len(articles) - 1:
                    print(f'\t+---{a}', file=output)
                    print('', file=output)
                else:
                    print(f'\t|---{a}', file=output)


def _xlsx_cluster_cell_border(size: int, x: int, a_cell: Cell):
    if x == 0 and size > 1:
        a_cell.border = Border(
            top=Side(style='thin'),
        )


def cluster_print_sheet(wb: Workbook, sheet_name: Optional[str], clusters: Dict[int, List[Article]]):
    if sheet_name:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.create_sheet()

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 38

    ws.append([
        # A        B        C            D            E       F         G          H
        'Cluster', 'Title', 'Published', 'Broadcast', 'Type', 'Source', 'Created', 'UUID'
    ])
    # Freeze the first row
    ws.freeze_panes = 'A2'

    for c, k in enumerate(clusters.keys()):
        articles: List[Article] = clusters[k]
        cl = articles[0]
        articles.sort(key=lambda article: article.created)
        size = len(articles)
        for x, a in enumerate(articles):
            broadcast = ''
            if a.mediaType['name'] == 'tv' or a.mediaType['name'] == 'radio':
                broadcast = a.published.replace(tzinfo=None)

            kl_token = os.environ.get('KMAP_TOKEN', None)
            previewUrl = ''
            pdfUrl = ''
            url = ''
            if kl_token:
                rel_path = os.path.join(
                    str(a.created.year), f"{a.created.month:02d}", f"{a.created.day:02d}", a.uuid
                )
                if a.url:
                    url = 'https://www.klipingmap.com/v3.0/media/redirect?filePath=' + rel_path
                    url += '&purpose=2&language=en&summaryType=override&showHighlights=true&&dcStringToken=' + kl_token

                previewUrl = 'https://www.klipingmap.com/v3.0/media/html?filePath=' + rel_path
                previewUrl += '&purpose=2&language=en&summaryType=override&showHighlights=true&dcStringToken=' + kl_token
                pdfUrl = 'https://www.klipingmap.com/v3.0/media/pdf?filePath=' + rel_path
                pdfUrl += '&purpose=2&language=en&summaryType=override&showHighlights=true&&dcStringToken=' + kl_token

            row = []
            a_cell: WriteOnlyCell = WriteOnlyCell(ws, value=cl.uuid)
            a_cell.style = 'uuid'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.title)
            if a.uuid == cl.uuid and size > 1:
                a_cell.style = 'hl_b'
            else:
                a_cell.style = 'hl'
            a_cell.hyperlink = previewUrl

            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.published.replace(tzinfo=None))
            a_cell.style = 'pub'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=broadcast)
            a_cell.style = 'br'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.mediaType['name'])
            a_cell.style = 'def'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.media)
            a_cell.style = 'def'
            if url:
                a_cell.style = 'hl'
                a_cell.hyperlink = url
            else:
                a_cell.style = 'def'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.created.replace(tzinfo=None))
            a_cell.style = 'cr'
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)

            a_cell = WriteOnlyCell(ws, value=a.uuid)
            a_cell.style = 'uuid'
            a_cell.hyperlink = pdfUrl
            _xlsx_cluster_cell_border(size, x, a_cell)
            row.append(a_cell)
            ws.append(row)


def cluster_create_wb():
    wb = Workbook(write_only=True)

    arial = Font(name='Arial', size=10)
    default_style = NamedStyle(name='def')
    default_style.font = arial
    wb.add_named_style(default_style)

    default_style_b = NamedStyle(name='def_b')
    default_style_b.font = Font(name='Arial', size=10, b=True)
    default_style_b.fill = PatternFill('solid', fgColor='FEE135')
    wb.add_named_style(default_style_b)

    link_style = NamedStyle(name='hl')
    link_style.font = Font(name='Arial', size=10, color='0000FF', underline='single')
    wb.add_named_style(link_style)

    link_style_b = NamedStyle(name='hl_b')
    link_style_b.font = Font(name='Arial', size=10, color='0000FF', underline='single', b=True)
    link_style_b.fill = PatternFill('solid', fgColor='FEE135')
    wb.add_named_style(link_style_b)

    uuid_style = NamedStyle(name='uuid')
    uuid_style.alignment = Alignment(horizontal='center')
    uuid_style.font = Font(name='Courier New', size=10)
    wb.add_named_style(uuid_style)

    published_style = NamedStyle(name='pub')
    published_style.number_format = 'dd. mm. yyyy'
    published_style.font = arial
    published_style.alignment = Alignment(horizontal='center')
    wb.add_named_style(published_style)

    broadcast_style = NamedStyle(name='br')
    broadcast_style.number_format = 'HH:mm'
    broadcast_style.font = arial
    broadcast_style.alignment = Alignment(horizontal='center')
    wb.add_named_style(broadcast_style)

    created_style = NamedStyle(name='cr')
    created_style.number_format = 'dd. mm. yyyy HH:mm:ss'
    created_style.font = arial
    created_style.alignment = Alignment(horizontal='center')
    wb.add_named_style(created_style)

    return wb


def cluster_print_xlsx(clusters: Dict[int, List[Article]], file_name: Optional[str] = None):
    wb = cluster_create_wb()
    cluster_print_sheet(wb, None, clusters)
    if file_name:
        wb.save(file_name)
    else:
        wb.save('clusters.xlsx')


def cluster_print_csv(clusters: Dict[int, List[Article]], file_name: Optional[str] = None):
    data = []
    for c, k in enumerate(clusters.keys()):
        articles: List[Article] = clusters[k]
        cluster_top = articles[0]
        # cluster_first = min(articles, key=attrgetter('created'))
        for x, a in enumerate(articles):

            broadcast = ''
            if a.mediaType['name'] == 'tv' or a.mediaType['name'] == 'radio':
                broadcast = a.published.strftime("%H:%M")

            kl_token = os.environ.get('KMAP_TOKEN', None)
            previewUrl = ''
            pdfUrl = ''
            url = ''
            if kl_token:
                rel_path = os.path.join(
                    str(a.created.year), f"{a.created.month:02d}", f"{a.created.day:02d}", a.uuid
                )
                if a.url:
                    url = 'https://stag-www.klipingmap.com/v3.0/media/redirect?filePath=' + rel_path
                    url += '&purpose=2&language=en&summaryType=override&showHighlights=true&&dcStringToken=' + kl_token

                previewUrl = 'https://www.klipingmap.com/v3.0/media/html?filePath=' + rel_path
                if a.mediaType['name'] == 'print':
                    previewUrl = 'https://www.klipingmap.com/v3.0/media/image?filePath=' + rel_path

                previewUrl += '&purpose=2&language=en&summaryType=override&showHighlights=true&&dcStringToken=' + kl_token
                pdfUrl = 'https://www.klipingmap.com/v3.0/media/pdf?filePath=' + rel_path
                pdfUrl += '&purpose=2&language=en&summaryType=override&showHighlights=true&&dcStringToken=' + kl_token

            data.append(
                (
                    cluster_top.uuid,
                    a.title,
                    a.published.strftime("%d. %m. %Y"),
                    broadcast,
                    a.mediaType['name'],
                    a.media,
                    a.published.strftime("%d. %m. %Y %H:%M:%S"),
                    a.uuid,
                    url,
                    previewUrl,
                    pdfUrl
                )
            )
    df = pd.DataFrame(data, columns=[
        'Cluster', 'Title', 'Published', 'Broadcast', 'Type', 'Source', 'Created', 'UUID', 'URL', 'Preview', 'PDF'
    ])
    df.to_csv(file_name, index=False)


def cluster_prep_json(clusters: Dict[int, List[Article]], country: str, start: datetime, end: datetime):
    data = {'from': start.isoformat(), 'to': end.isoformat()}
    if country:
        data['country'] = country
    data['clusters'] = []
    for c, k in enumerate(clusters.keys()):
        articles: List[Article] = clusters[k]
        cl = articles[0]
        size = len(articles)
        cluster = {'uuid': cl.uuid, 'size': size, 'idx': c, 'title': cl.title, 'articles': []}
        articles.sort(key=lambda article: article.created)
        data['clusters'].append(cluster)
        for x, a in enumerate(articles):
            cl_article = {
                'uuid': a.uuid,
                'published': a.published.isoformat(),
                'crated': a.created.isoformat(),
                'media': a.media,
                'rubric': a.rubric,
                'language': a.language,
                'country': a.country['name'],
                'mediaType': a.mediaType['name'],
                'mediaReach': a.mediaReach,
                'title': a.title,
                'relPath': a.data['relPath']
            }
            if a.url:
                cl_article['url'] = a.url
            cl_article['body'] = a.body
            cluster['articles'].append(cl_article)
    return data


def compare_clusterings(a_first, a_second, cf_name='First Clustering', cs_name='Second Clustering'):
    diff = list(set(a_first) - set(a_second))
    if diff:
        print(f'{cf_name}[{len(a_first)}] {len(diff)} cluster articles are missing in {cs_name}[{len(a_second)}]')
        for x, a in enumerate(diff):
            if x == len(diff) - 1:
                print(f'\t+---{a}')
            else:
                print(f'\t|---{a}')
        return -1
    else:
        delta = len(a_first) - len(a_second)
        if delta == 0:
            print(f'Clusters are the same!')
        else:
            print(f'{cs_name} cluster [{a_first[0].title}] has all articles that are in {cf_name}!')
    return delta


def clusters_compare(c_first, c_second, fc_name, sc_name):

    for k in c_first.keys():
        if k not in c_second:
            print(f'{sc_name} is missing cluster [{k}]')
            continue

        a_first: List[Article] = c_first[k]
        a_second: List[Article] = c_second[k]

        if len(a_first) < 2 and len(a_second) < 2:
            continue

        print(f'Compare cluster [{a_first[0].title}] uuid [{k}]')
        ret = compare_clusterings(a_first, a_second, fc_name, sc_name)
        if ret == 0:
            print('---------------------------------------------------------------------------------------------------'
                  '---------------------------------------------------------------------------------------------------')
            print('')
            continue
        compare_clusterings(a_second, a_first, sc_name, fc_name)
        print('---------------------------------------------------------------------------------------------------'
              '---------------------------------------------------------------------------------------------------')
        print('')
